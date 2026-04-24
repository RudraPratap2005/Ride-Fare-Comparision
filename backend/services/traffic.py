from functools import lru_cache
import os
from pathlib import Path

TRAFFIC_PROFILE_MULTIPLIERS = {
    "Low": 0.98,
    "Medium": 1.03,
    "High": 1.09,
}


def infer_default_traffic_level(trip_hour: int) -> str:
    if (8 <= trip_hour <= 11) or (18 <= trip_hour <= 21):
        return "High"
    if (
        (7 <= trip_hour < 8)
        or (11 < trip_hour <= 13)
        or (17 <= trip_hour < 18)
        or (21 < trip_hour <= 22)
    ):
        return "Medium"
    return "Low"


def _candidate_dataset_paths() -> list[Path]:
    user_home = Path.home()
    configured = os.getenv("RIDESCOPE_TRAFFIC_DATASET_PATH")

    candidates = []
    if configured:
        candidates.append(Path(configured))

    candidates.extend(
        [
            Path("data/Bookings-100000-Rows.xlsx"),
            user_home / "Downloads" / "Bookings-100000-Rows.xlsx",
        ]
    )
    return candidates


@lru_cache(maxsize=1)
def get_dataset_hour_profile() -> dict | None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    dataset_path = next((path for path in _candidate_dataset_paths() if path.exists()), None)
    if dataset_path is None:
        return None

    workbook = load_workbook(dataset_path, read_only=True, data_only=True)
    sheet_name = "July" if "July" in workbook.sheetnames else workbook.sheetnames[0]
    worksheet = workbook[sheet_name]

    hour_counts = {hour: 0 for hour in range(24)}
    total_rows = 0

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        date_value = row[0]
        time_value = row[1]
        hour = None

        if hasattr(time_value, "hour"):
            hour = time_value.hour
        elif hasattr(date_value, "hour"):
            hour = date_value.hour

        if hour is None:
            continue

        hour_counts[hour] += 1
        total_rows += 1

    if total_rows == 0:
        return None

    counts = list(hour_counts.values())
    sorted_counts = sorted(counts)

    return {
        "source": str(dataset_path),
        "sheet_name": sheet_name,
        "total_rows": total_rows,
        "hour_counts": hour_counts,
        "average_count": sum(counts) / len(counts),
        "high_threshold": max(sorted_counts[int(len(sorted_counts) * 0.75)], 1),
        "low_threshold": sorted_counts[int(len(sorted_counts) * 0.25)],
    }


def infer_dataset_traffic_level(trip_hour: int) -> str | None:
    profile = get_dataset_hour_profile()
    if profile is None:
        return None

    count = profile["hour_counts"][trip_hour]
    average_count = profile["average_count"]
    high_threshold = max(profile["high_threshold"], average_count * 1.02)
    low_threshold = min(profile["low_threshold"], average_count * 0.985)

    if count >= high_threshold:
        return "High"
    if count <= low_threshold:
        return "Low"
    return "Medium"


def infer_traffic_level(trip_hour: int) -> str:
    default_level = infer_default_traffic_level(trip_hour)
    dataset_level = infer_dataset_traffic_level(trip_hour)

    if dataset_level is None:
        return default_level

    if dataset_level == "High":
        if default_level == "Low":
            return "Medium"
        return "High"

    if dataset_level == "Low":
        if default_level == "High":
            return "Medium"
        return "Low"

    return default_level


def is_peak_hour(trip_hour: int) -> bool:
    return infer_traffic_level(trip_hour) == "High"


def get_route_traffic_multiplier(distance_km: float, duration_min: float) -> float:
    if distance_km <= 0 or duration_min <= 0:
        return 1.0

    avg_speed_kmph = distance_km / (duration_min / 60)

    if avg_speed_kmph < 12:
        return 1.16
    if avg_speed_kmph < 18:
        return 1.10
    if avg_speed_kmph < 25:
        return 1.05
    return 1.0
